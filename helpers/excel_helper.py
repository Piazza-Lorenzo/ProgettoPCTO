import openpyxl
import os

def create_excel_if_not_exists():
    """
    Crea un file Excel con la struttura necessaria se non esiste.

    Returns:
        str: Il nome del file Excel
    """
    excel_filename = "lista_aziende.xlsx"
    if not os.path.exists(excel_filename):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Nome Azienda'
        ws['B1'] = 'URL'
        ws['C1'] = 'Email'
        ws['D1'] = 'Telefono'
        ws['E1'] = 'Settore'
        wb.save(excel_filename)
        print(f"✓ File Excel creato: {excel_filename}")
    else:
        print(f"✓ File Excel esistente trovato: {excel_filename}")
    return excel_filename

def load_existing_companies(excel_filename):
    """
    Carica le aziende esistenti dal file Excel.

    Args:
        excel_filename (str): Il nome del file Excel

    Returns:
        list: Lista di aziende esistenti
    """
    if not os.path.exists(excel_filename):
        return []
    
    wb = openpyxl.load_workbook(excel_filename)
    ws = wb.active
    companies = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # Salta l'intestazione
        if row[0]:  # Se c'è un nome
            companies.append({
                'name': row[0],
                'url': row[1],
                'email': row[2],
                'phone': row[3],
                'sector': row[4] if len(row) > 4 else None
            })
    return companies

def add_company_to_excel(excel_filename, company, sector):
    """
    Aggiunge una singola azienda al file Excel.

    Args:
        excel_filename (str): Il nome del file Excel
        company (dict): I dati dell'azienda da aggiungere
        sector (str): Il settore dell'azienda
    """
    wb = openpyxl.load_workbook(excel_filename)
    ws = wb.active
    ws.append([company['name'], company['url'], company['email'], company['phone'], sector])
    wb.save(excel_filename)
    print(f"✓ Azienda '{company['name']}' aggiunta al file Excel")

def save_excel_file(companies, sector):
    """
    Salva i dati delle aziende in un file Excel.

    Args:
        companies (list): Lista di aziende da salvare
        sector (str): Il settore per il quale sono state trovate aziende
    """
    # Crea una nuova cartella di lavoro Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = 'Nome Azienda'
    ws['B1'] = 'URL'
    ws['C1'] = 'Email'
    ws['D1'] = 'Telefono'
    ws['E1'] = 'Settore'

    # Aggiungi i dati al file Excel
    for company in companies:
        ws.append([company['name'], company['url'], company['email'], company['phone'], sector])

    # Salva il file Excel
    excel_filename = "lista_aziende.xlsx"
    wb.save(excel_filename)
    print(f"✓ File Excel salvato: {excel_filename}")
    return excel_filename